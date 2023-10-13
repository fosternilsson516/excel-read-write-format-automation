# excel_formatting.py

from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def apply_formatting(worksheet):
    arial_font = Font(name='Arial', size=10)
    light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True)
    top_left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    middle_center_alignment = Alignment(horizontal='center', vertical='center')

    worksheet.row_dimensions[1].height = 20

    for row in worksheet.iter_rows():
        if row[0].row != 1:  # Exclude header row from height adjustment
            worksheet.row_dimensions[row[0].row].height = 310
    
        for cell in row:
            cell.font = arial_font

            if cell.row != 1:  
            # Default word wrap
                cell.alignment = wrap_alignment
        # Specific alignments based on column
            col_letter = cell.column_letter
            if col_letter in ['A', 'B', 'E', 'F', 'G', 'H']:
                cell.alignment = top_left_alignment
            elif col_letter in ['C', 'D']:
                cell.alignment = middle_center_alignment
            else:
                cell.alignment = wrap_alignment    

        # Fill cells with content, excluding the header, with light green
            if cell.row != 1 and cell.value:  # Exclude header row and ensure cell has content
                cell.fill = light_green_fill

# Adding 3 white spaces (rows) between each entry
    for i in reversed(range(3, worksheet.max_row + 1)):  # Start from the row after the first entry
        worksheet.insert_rows(i, 3)
    # Auto-resize columns
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(29, max_length + 2)  # Adjusting but ensuring it doesn't exceed 25
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    # Add border to cells
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border

    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center') 
    for cell in worksheet["1:1"]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for cell in worksheet['G']:
        if cell.row != 1:  # Exclude header row
            cell.fill = light_green_fill
            cell.alignment = top_left_alignment
            cell.font = arial_font

    for cell in worksheet['H']:
        if cell.row != 1:  # Exclude header row
            cell.fill = light_green_fill
            cell.alignment = top_left_alignment
            cell.font = arial_font
    return worksheet